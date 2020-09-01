from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox


class Library(object):
    def __init__(self, available_books, students_list, admin_info):
        self.available_books = available_books
        self.students_list = students_list
        self.admin_info = admin_info

    def get_admin_info(self):
        return self.admin_info

    def set_admin_info(self, admin_info):
        self.admin_info = admin_info

    def get_available_books(self):
        return self.available_books

    def get_students_list(self):
        return self.students_list

    def display_available_books(self):
        print("The books currently available in our library are:")
        for book in self.available_books:
            print(book)
        print(50 * "-")

    def add_book(self, id, title, author, stock="", rack_no=""):
        self.available_books.append(Book(id, title, author, stock, rack_no))

    def remove_book(self, id):
        for book in self.available_books:
            if book.get_id() == id:
                self.available_books.remove(book)
                break

    def update_book(self, title_old, id, title, author, stock="", rack_no=""):
        for book in self.available_books:
            if book.get_title() == title_old:
                book.update_book(title, author, id, stock, rack_no)

    def search(self, title):
        matched_books = []

        for book in LMS.available_books:
            if title.lower() in book.get_title().lower() or title.lower() == book.get_title().lower():
                matched_books.append(book)

        return matched_books

    def get_book_info(self, title):
        book_found = False
        for book in self.available_books:
            if book.get_title().lower() == title.lower():
                book.get_info()
                book_found = True
                break
        if not book_found:
            print("Book could not be found.")

    def lend_book(self, title, student_id):

        pass

    def return_book(self, title, student_id):

        pass

    def get_student_info(self, id):
        for student in self.students_list:
            if student.get_id() == id:
                student.get_info()
                break

    def add_student(self, name, id):
        self.students_list.append(Student(name, id))

    def remove_student(self, id):
        for student in self.students_list:
            if student.get_id() == id:
                self.students_list.remove(student)

    def get_late_fee(self, borrow_date, return_date):
        d1_list = borrow_date.split('/')
        d2_list = return_date.split('/')

        d1 = date(int(d1_list[2]), int(d1_list[1]), int(d1_list[0]))
        d2 = date(int(d2_list[2]), int(d2_list[1]), int(d2_list[0]))

        delta = d2 - d1
        days = delta.days
        days_late = days - 7
        late_fee = 2 * days_late

        if late_fee > 500:
            return 500
        else:
            return late_fee

    def display_students(self):
        for student in self.students_list:
            print(student)
        print(50 * "-")


class Book(object):
    def __init__(self, id, title, author, stock="1", rack_no=""):
        self.id = str(id)
        self.title = title
        self.author = author
        self.stock = stock
        self.rack_no = rack_no

    def __str__(self):
        separator = 50 * "-"
        return f"{separator}\nID:\t{self.id}\nTitle:\t{self.title}\nAuthor:\t{self.author}"

    def set_title(self, title):
        self.title = title

    def get_title(self):
        return self.title

    def set_author(self, author):
        self.author = author

    def get_author(self):
        return self.author

    def get_id(self):
        return self.id

    def set_id(self, id):
        self.id = id

    def add_stock(self, n):
        self.stock = str(int(self.stock) + n)

    def get_stock(self):
        return self.stock

    def set_rack_no(self, rack_no):
        self.rack_no = rack_no

    def get_rack_no(self):
        return self.rack_no

    def get_info(self):
        return f"ID:\t{self.id}\nTitle:\t{self.title}\nAuthor:\t{self.author}\nStock:\t{self.stock}\nRack No:\t{self.rack_no}"

    def update_book(self, title, author, id, stock="1", rack_no="B12"):
        self.title = title
        self.author = author
        self.id = id
        self.stock = stock
        self.rack_no = rack_no


class Student(object):
    def __init__(self, name, id, borrowed_books=[]):
        self.name = name
        self.id = id
        self.borrowed_books = borrowed_books

    def __str__(self):
        separator = 50 * "-"
        return f"{separator}\nName:\t{self.name}\nID:\t\t{self.id}"

    def get_id(self):
        return self.id

    def get_name(self):
        return self.name

    def has_passed_book_limit(self):
        if len(self.borrowed_books) >= 3:
            return True
        else:
            return False

    def get_borrowed_books_list(self):

        list = []
        for book in self.borrowed_books:
            list.append(book[0])

        return list

    def get_borrowed_books(self):
        text = ""
        for list in self.borrowed_books:
            text += f"{list[0]}, "

        return text

    def get_borrow_date(self, book_title):
        for book in self.borrowed_books:
            if book[0].lower() == book_title.lower():
                return book[1]

    def get_borrow_dates(self):
        text = ""
        for list in self.borrowed_books:
            text += f"{list[1]}, "

        return text

    def get_info(self):
        text = ""
        text += f"Name:\t{self.name}\nID:\t{self.id}\nBooks:\n\n"
        for book in self.borrowed_books:
            text += f"{book[0]}\n{book[1]}\n\n"
        return text

    def borrow_book(self, book_title, borrow_date):
        self.borrowed_books.append([book_title, borrow_date])

    def return_book(self, book):
        for borrowed_book in self.borrowed_books:
            if borrowed_book[0] == book:
                self.borrowed_books.remove(borrowed_book)


def check_return_date_validity(borrow_date, return_date):
    borrow_date = borrow_date.split('/')
    return_date = return_date.split('/')

    borrow_date = date(int(borrow_date[2]), int(borrow_date[1]), int(borrow_date[0]))
    return_date = date(int(return_date[2]), int(return_date[1]), int(return_date[0]))

    if borrow_date.year > return_date.year:
        return False
    elif borrow_date.year == return_date.year and borrow_date.month > return_date.month:
        return False
    elif borrow_date.year == return_date.year and borrow_date.month == return_date.month and borrow_date.day > return_date.day:
        return False
    else:
        return True



def check_credentials(credentials):
    username = credentials[0]
    password = credentials[1]

    if [username, password] == LMS.get_admin_info():
        return True
    else:
        return False


def books_max_row_style():
    for i in range(1, 6):
        books_db.row_dimensions[books_db.max_row].height = 20
        books_db.cell(row=books_db.max_row, column=i).alignment = Alignment(horizontal="center", vertical="center")
        books_db.cell(row=books_db.max_row, column=i).fill = PatternFill(start_color="9bc2e6", fill_type="solid")


def students_max_row_style():
    for i in range(1, 5):
        students_db.row_dimensions[students_db.max_row].height = 20
        students_db.cell(row=students_db.max_row, column=i).alignment = Alignment(horizontal="center", vertical="center")
        students_db.cell(row=students_db.max_row, column=i).fill = PatternFill(start_color="9bc2e6", fill_type="solid")


def max_row_add_book(book):
    row_list = list(book.__dict__.values())
    books_db.append(row_list)
    books_max_row_style()
    wb.save("Database.xlsx")


def max_row_add_student(student):
    row = students_db.max_row + 1
    students_db.cell(row=row, column=1).value = student.get_id()
    students_db.cell(row=row, column=2).value = student.get_name()
    students_db.cell(row=row, column=3).value = student.get_borrowed_books()
    students_db.cell(row=row, column=4).value = student.get_borrow_dates()

    students_max_row_style()
    wb.save("Database.xlsx")


def remove_book_from_db(book_id):
    column = 1
    row = 2
    while True:
        if str(books_db.cell(row=row, column=column).value) == book_id:
            books_db.delete_rows(row)
            break
        else:
            row += 1
    wb.save("Database.xlsx")


def remove_student_from_db(student_id):
    column = 1
    row = 2
    while True:
        if str(students_db.cell(row=row, column=column).value).lower() == student_id.lower():
            students_db.delete_rows(row)
            break
        else:
            row += 1
    wb.save("Database.xlsx")


def edit_book_in_db(book, title_old):
    column = 2
    row = 2
    while True:
        if str(books_db.cell(row=row, column=column).value) == title_old:
            books_db.cell(row=row, column=1).value = book.get_id()
            books_db.cell(row=row, column=2).value = book.get_title()
            books_db.cell(row=row, column=3).value = book.get_author()
            books_db.cell(row=row, column=4).value = book.get_stock()
            books_db.cell(row=row, column=5).value = book.get_rack_no()
            break
        else:
            row += 1
    wb.save("Database.xlsx")


def get_books_from_db():
    books = []

    for row_n in range(2, books_db.max_row + 1):
        id = books_db.cell(row=row_n, column=1).value
        title = books_db.cell(row=row_n, column=2).value
        author = books_db.cell(row=row_n, column=3).value
        stock = books_db.cell(row=row_n, column=4).value
        rack_no = books_db.cell(row=row_n, column=5).value

        books.append(Book(id, title, author, stock, rack_no))
    return books


def get_students_from_db():
    students = []

    for row_n in range(2, students_db.max_row + 1):
        id = students_db.cell(row=row_n, column=1).value
        name = students_db.cell(row=row_n, column=2).value
        borrowed_books_titles = students_db.cell(row=row_n, column=3).value
        if borrowed_books_titles is None:
            borrowed_books_list = []
        else:
            borrow_dates = students_db.cell(row=row_n, column=4).value
            borrowed_books_titles_list = borrowed_books_titles.split(", ")
            borrow_dates_list = borrow_dates.split(", ")

            for title in borrowed_books_titles_list:
                if title == "":
                    borrowed_books_titles_list.remove(title)

            for date in borrow_dates_list:
                if date == "":
                    borrow_dates_list.remove(date)

            borrowed_books_list = []
            for i in range(len(borrowed_books_titles_list)):
                borrowed_books_list.append([borrowed_books_titles_list[i], borrow_dates_list[i]])

        students.append(Student(name, id, borrowed_books_list))

    return students


def get_admin_info_from_db():
    admin_info = []
    admin_info.append(admin_db.cell(row=1, column=2).value)
    admin_info.append(admin_db.cell(row=2, column=2).value)
    return admin_info


def set_admin_info_in_db(credentials):
    admin_db.cell(row=1, column=2).value = credentials[0]
    admin_db.cell(row=2, column=2).value = credentials[1]
    wb.save("Database.xlsx")


class LoginScreen(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(624, 400)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.login_menu_label = QtWidgets.QLabel(self.centralwidget)
        self.login_menu_label.setGeometry(QtCore.QRect(150, 10, 380, 41))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(18)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.login_menu_label.setFont(font)
        self.login_menu_label.setObjectName("login_menu_label")
        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(20, 50, 581, 321))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(14)
        self.groupBox.setFont(font)
        self.groupBox.setTitle("")
        self.groupBox.setObjectName("groupBox")
        self.admin_label = QtWidgets.QLabel(self.groupBox)
        self.admin_label.setGeometry(QtCore.QRect(60, 40, 181, 31))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.admin_label.setFont(font)
        self.admin_label.setObjectName("admin_label")
        self.student_label = QtWidgets.QLabel(self.groupBox)
        self.student_label.setGeometry(QtCore.QRect(360, 40, 131, 31))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.student_label.setFont(font)
        self.student_label.setObjectName("student_label")
        self.username_label = QtWidgets.QLabel(self.groupBox)
        self.username_label.setGeometry(QtCore.QRect(30, 105, 81, 16))
        self.username_label.setObjectName("username_label")
        self.password_label = QtWidgets.QLabel(self.groupBox)
        self.password_label.setGeometry(QtCore.QRect(30, 135, 81, 16))
        self.password_label.setObjectName("password_label")
        self.username_input = QtWidgets.QLineEdit(self.groupBox)
        self.username_input.setGeometry(QtCore.QRect(130, 110, 113, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.username_input.setFont(font)
        self.username_input.setObjectName("username_input")
        self.password_input = QtWidgets.QLineEdit(self.groupBox)
        self.password_input.setGeometry(QtCore.QRect(130, 140, 113, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.password_input.setFont(font)
        self.password_input.setObjectName("password_input")
        self.admin_login_b = QtWidgets.QPushButton(self.groupBox)
        self.admin_login_b.setGeometry(QtCore.QRect(70, 210, 151, 51))
        self.admin_login_b.setObjectName("admin_login_b")
        self.admin_login_b.setStatusTip("Login as an Admin")
        self.student_login_b = QtWidgets.QPushButton(self.groupBox)
        self.student_login_b.setGeometry(QtCore.QRect(360, 210, 151, 51))
        self.student_login_b.setObjectName("student_login_b")
        self.student_login_b.setStatusTip("Login as a Student")
        self.line = QtWidgets.QFrame(self.groupBox)
        self.line.setGeometry(QtCore.QRect(290, 30, 20, 270))
        self.line.setFrameShape(QtWidgets.QFrame.VLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 624, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.admin_login_b.clicked.connect(self.admin_login)
        self.student_login_b.clicked.connect(self.student_login)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Login"))
        self.login_menu_label.setText(_translate("MainWindow", "Library Management System"))
        self.admin_label.setText(_translate("MainWindow", "Administrator Login"))
        self.student_label.setText(_translate("MainWindow", "Student Login"))
        self.username_label.setText(_translate("MainWindow", "Username"))
        self.password_label.setText(_translate("MainWindow", "Password"))
        self.admin_login_b.setText(_translate("MainWindow", "Admin Login"))
        self.student_login_b.setText(_translate("MainWindow", "Student Login"))
        self.admin_label.adjustSize()
        self.student_label.adjustSize()
        self.username_label.adjustSize()
        self.password_label.adjustSize()

    def admin_login_popup(self, condition):
        msg = QMessageBox()
        msg.setWindowTitle("Admin Login")

        if condition == "success":
            msg.setText("Login is Successful.")
        else:
            msg.setText("Login Failed.       ")
            msg.setIcon(QMessageBox.Critical)

        x = msg.exec_()

    def admin_login(self):
        t = self.username_input.text()
        username=t.replace(" ","")
        password = self.password_input.text()
        credentials = [username, password]
        if check_credentials(credentials):
            self.admin_login_popup("success")
            self.username_input.clear()
            self.password_input.clear()
            login_screen.close()
            admin_menu.show()

        else:
            self.admin_login_popup("failed")

    def student_login(self):
        login_screen.close()
        students_menu.show()


class StudentsMenu(object):
    def setupUi(self, StudentsWindow):
        StudentsWindow.setObjectName("MainWindow")
        StudentsWindow.resize(634, 536)
        self.centralwidget = QtWidgets.QWidget(StudentsWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.student_menu_label = QtWidgets.QLabel(self.centralwidget)
        self.student_menu_label.setGeometry(QtCore.QRect(230, 10, 161, 41))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(18)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.student_menu_label.setFont(font)
        self.student_menu_label.setObjectName("student_menu_label")
        self.search_groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.search_groupBox.setGeometry(QtCore.QRect(335, 80, 251, 131))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(10)
        self.search_groupBox.setFont(font)
        self.search_groupBox.setObjectName("search_groupBox")
        self.search_label = QtWidgets.QLabel(self.search_groupBox)
        self.search_label.setGeometry(QtCore.QRect(5, 40, 81, 21))
        self.search_label.setObjectName("search_label")
        self.search_title = QtWidgets.QLineEdit(self.search_groupBox)
        self.search_title.setGeometry(QtCore.QRect(90, 40, 141, 21))
        self.search_title.setObjectName("search_title")
        self.search_b = QtWidgets.QPushButton(self.search_groupBox)
        self.search_b.setGeometry(QtCore.QRect(130, 80, 91, 31))
        self.search_b.setObjectName("search_b")
        self.viewAll_b = QtWidgets.QPushButton(self.search_groupBox)
        self.viewAll_b.setGeometry(QtCore.QRect(30, 80, 91, 31))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(12)
        self.viewAll_b.setFont(font)
        self.viewAll_b.setObjectName("viewAll_b")
        self.bookDescription_textBrowser = QtWidgets.QTextBrowser(self.centralwidget)
        self.bookDescription_textBrowser.setGeometry(QtCore.QRect(320, 270, 290, 191))
        self.bookDescription_textBrowser.setObjectName("bookDescription_textBrowser")
        self.bookDescription_label = QtWidgets.QLabel(self.centralwidget)
        self.bookDescription_label.setGeometry(QtCore.QRect(320, 240, 160, 21))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.bookDescription_label.setFont(font)
        self.bookDescription_label.setObjectName("bookDescription_label")
        self.booksList_label = QtWidgets.QLabel(self.centralwidget)
        self.booksList_label.setGeometry(QtCore.QRect(30, 90, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.booksList_label.setFont(font)
        self.booksList_label.setObjectName("booksList_label")
        self.books_list = QtWidgets.QListWidget(self.centralwidget)
        self.books_list.setGeometry(QtCore.QRect(30, 120, 256, 341))
        self.books_list.setObjectName("books_list")
        StudentsWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(StudentsWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 634, 21))
        self.menubar.setObjectName("menubar")
        self.menubar.setObjectName("menubar")
        self.menuFile = QtWidgets.QMenu(self.menubar)
        self.menuFile.setObjectName("menuFile")
        self.menuHelp = QtWidgets.QMenu(self.menubar)
        self.menuHelp.setObjectName("menuHelp")
        self.menuThemes = QtWidgets.QMenu(self.menubar)
        self.menuThemes.setObjectName("menuThemes")
        StudentsWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(StudentsWindow)
        self.statusbar.setObjectName("statusbar")
        StudentsWindow.setStatusBar(self.statusbar)
        self.actionLogout = QtWidgets.QAction(StudentsWindow)
        self.actionLogout.setObjectName("actionLogout")
        self.actionExit = QtWidgets.QAction(StudentsWindow)
        self.actionExit.setObjectName("actionExit")
        self.actionBlue = QtWidgets.QAction(StudentsWindow)
        self.actionBlue.setObjectName("actionBlue")
        self.actionCoral = QtWidgets.QAction(StudentsWindow)
        self.actionCoral.setObjectName("actionCoral")
        self.actionAbout = QtWidgets.QAction(StudentsWindow)
        self.actionAbout.setObjectName("actionAbout")
        self.actionBasic = QtWidgets.QAction(StudentsWindow)
        self.actionBasic.setObjectName("actionBasic")
        self.actionAqua = QtWidgets.QAction(StudentsWindow)
        self.actionAqua.setObjectName("actionAqua")
        self.actionCoral = QtWidgets.QAction(StudentsWindow)
        self.actionCoral.setObjectName("actionCoral")
        self.menuFile.addAction(self.actionLogout)
        self.menuFile.addSeparator()
        self.menuFile.addAction(self.actionExit)
        self.menuHelp.addAction(self.actionAbout)
        self.menuThemes.addAction(self.actionBasic)
        self.menuThemes.addAction(self.actionAqua)
        self.menuThemes.addAction(self.actionCoral)
        self.menubar.addAction(self.menuFile.menuAction())
        self.menubar.addAction(self.menuThemes.menuAction())
        self.menubar.addAction(self.menuHelp.menuAction())

        self.actionCoral.triggered.connect(self.change_to_coral)
        self.actionBasic.triggered.connect(self.change_back_to_basic)
        self.actionAqua.triggered.connect(self.change_to_blue)
        self.viewAll_b.clicked.connect(self.view_all)
        self.books_list.clicked.connect(self.show_book_description)
        self.search_b.clicked.connect(self.search)
        self.actionLogout.triggered.connect(self.logout)
        self.actionExit.triggered.connect(self.exit)

        self.actionAbout.triggered.connect(self.about)

        self.retranslateUi(StudentsWindow)
        QtCore.QMetaObject.connectSlotsByName(StudentsWindow)

    def retranslateUi(self, StudentsWindow):
        _translate = QtCore.QCoreApplication.translate
        StudentsWindow.setWindowTitle(_translate("StudentsWindow", "Library's Book Collection"))
        self.student_menu_label.setText(_translate("StudentsWindow", "Students Menu"))
        self.search_groupBox.setTitle(_translate("StudentsWindow", "Search"))
        self.search_label.setText(_translate("StudentsWindow", "Book Title:"))
        self.search_b.setStatusTip(_translate("StudentsWindow", "Search for a Book by Title in the Library\'s Collection"))
        self.search_b.setText(_translate("StudentsWindow", "Search"))
        self.viewAll_b.setStatusTip(_translate("StudentsWindow", "Display all Books in the Library\'s Collection"))
        self.viewAll_b.setText(_translate("StudentsWindow", "View All"))
        self.bookDescription_label.setText(_translate("StudentsWindow", "Book\'s Description"))
        self.booksList_label.setText(_translate("StudentsWindow", "Books List"))
        self.menuFile.setTitle(_translate("MainWindow", "File"))
        self.menuHelp.setTitle(_translate("MainWindow", "Help"))
        self.menuThemes.setTitle(_translate("MainWindow", "Themes"))
        self.actionLogout.setText(_translate("MainWindow", "Logout"))
        self.actionLogout.setStatusTip(_translate("MainWindow", "Go back to Main Menu"))
        self.actionLogout.setShortcut(_translate("MainWindow", "Ctrl+Q"))
        self.actionExit.setText(_translate("MainWindow", "Exit"))
        self.actionExit.setStatusTip(_translate("MainWindow", "Exit System"))
        self.actionExit.setShortcut(_translate("MainWindow", "Esc"))
        self.actionBlue.setText(_translate("MainWindow", "Basic"))
        self.actionCoral.setText(_translate("MainWindow", "Coral"))
        self.actionAbout.setText(_translate("MainWindow", "About"))
        self.actionAbout.setStatusTip(_translate("MainWindow", "About the System"))
        self.actionAbout.setShortcut(_translate("MainWindow", "Ctrl+H"))
        self.actionBasic.setText(_translate("MainWindow", "Basic"))
        self.actionBasic.setStatusTip(_translate("MainWindow", "Basic Theme"))
        self.actionAqua.setText(_translate("MainWindow", "Aqua"))
        self.actionAqua.setStatusTip(_translate("MainWindow", "Blue-ish Theme"))
        self.actionCoral.setText(_translate("MainWindow", "Coral"))
        self.actionCoral.setText(_translate("MainWindow", "Coral"))
        self.actionCoral.setStatusTip(_translate("MainWindow", "Pink-ish Theme"))
        self.student_menu_label.adjustSize()
        self.bookDescription_label.adjustSize()
        self.booksList_label.adjustSize()

    def view_all(self):
        self.search_title.clear()
        self.bookDescription_textBrowser.clear()
        self.books_list.clear()
        for book in LMS.get_available_books():
            self.books_list.addItem(book.get_title())

    def show_book_description(self):
        title = self.books_list.currentItem().text()

        for book in LMS.get_available_books():
            if title == book.get_title():

                text = book.get_info()

        self.bookDescription_textBrowser.setText(text)

    def search_popup(self, message):
        msg = QMessageBox()
        msg.setWindowTitle("Book Search")
        msg.setText(message)
        msg.setIcon(QMessageBox.Critical)

        x = msg.exec_()

    def search(self):
        self.bookDescription_textBrowser.clear()
        search_title = self.search_title.text()

        if search_title == "":
            self.search_popup("Please Enter a Title.")
        else:
            books_found = LMS.search(search_title)

            if len(books_found) == 0:
                self.search_popup("Book isn't Available.")

            else:
                self.books_list.clear()
                for book in books_found:
                    self.books_list.addItem(book.get_title())

    def about(self):
        msg = QMessageBox()
        msg.setWindowTitle("About")
        msg.setText("""The Library Management System LMS allows students to:
        1. View the Library's entire Book Collection.
        2. Search for any Book by Title.
        
Books can be borrowed for a period of one week, after which, the student must pay a late fee. 
        
        """)

        x = msg.exec_()

    def logout(self):
        students_menu.close()
        login_screen.show()

    def exit(self):
        students_menu.close()

    def change_back_to_basic(self):
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0,0,0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0,0,0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        students_menu.setPalette(palette)

    def change_to_coral(self):
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 85, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(183, 255, 251))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 170, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Highlight, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.HighlightedText, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.AlternateBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 85, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(183, 255, 251))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 170, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Highlight, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.HighlightedText, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.AlternateBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 120, 215))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Highlight, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.HighlightedText, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.AlternateBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.PlaceholderText, brush)
        students_menu.setPalette(palette)

    def change_to_blue(self):
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 85, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 85, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 255, 253))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(127, 49, 191, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 85, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 85, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 255, 253))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(127, 49, 191, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 0, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 0, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 0, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 255, 253))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 255, 253))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.PlaceholderText, brush)
        students_menu.setPalette(palette)


class ChangeCredsWindow(object):
    def setupUi(self, ChangeCredsWindow):
        ChangeCredsWindow.setObjectName("ChangeCredsWindow")
        ChangeCredsWindow.resize(474, 285)
        self.centralwidget = QtWidgets.QWidget(ChangeCredsWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(110, 30, 251, 191))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(10)
        self.groupBox.setFont(font)
        self.groupBox.setObjectName("groupBox")
        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setGeometry(QtCore.QRect(20, 70, 71, 16))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.groupBox)
        self.label_2.setGeometry(QtCore.QRect(20, 100, 71, 21))
        self.label_2.setObjectName("label_2")
        self.username_input = QtWidgets.QLineEdit(self.groupBox)
        self.username_input.setGeometry(QtCore.QRect(110, 70, 113, 20))
        self.username_input.setObjectName("username_input")
        self.password_input = QtWidgets.QLineEdit(self.groupBox)
        self.password_input.setGeometry(QtCore.QRect(110, 100, 113, 20))
        self.password_input.setObjectName("password_input")
        self.confirm_b = QtWidgets.QPushButton(self.groupBox)
        self.confirm_b.setGeometry(QtCore.QRect(70, 140, 101, 31))
        self.confirm_b.setObjectName("confirm_b")
        ChangeCredsWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(ChangeCredsWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 474, 21))
        self.menubar.setObjectName("menubar")
        ChangeCredsWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(ChangeCredsWindow)
        self.statusbar.setObjectName("statusbar")
        ChangeCredsWindow.setStatusBar(self.statusbar)

        self.confirm_b.clicked.connect(self.update_creds)

        self.retranslateUi(ChangeCredsWindow)
        QtCore.QMetaObject.connectSlotsByName(ChangeCredsWindow)

    def retranslateUi(self, ChangeCredsWindow):
        _translate = QtCore.QCoreApplication.translate
        ChangeCredsWindow.setWindowTitle(_translate("ChangeCredsWindow", "Admin Credentials"))
        self.groupBox.setTitle(_translate("ChangeCredsWindow", "Change Admin Credentials"))
        self.label.setText(_translate("ChangeCredsWindow", "Username"))
        self.label_2.setText(_translate("ChangeCredsWindow", "Password"))
        self.confirm_b.setText(_translate("ChangeCredsWindow", "Confirm"))
        self.label.adjustSize()
        self.label_2.adjustSize()

    def change_creds_popup(self, condition, message):
        msg = QMessageBox()
        msg.setWindowTitle("Change Admin Credentials")
        msg.setText(message)
        if message == "fail":
            msg.setIcon(QMessageBox.Critical)

        x = msg.exec_()

    def update_creds(self):
        username = self.username_input.text()
        password = self.password_input.text()
        if username == "" or password == "":
            self.change_creds_popup("fail", "Invalid input.\nPlease enter the new username and password.")
        else:
            LMS.set_admin_info([username, password])
            self.change_creds_popup("success", "Credentials changed Successfully.")
            set_admin_info_in_db([username, password])
            change_creds.close()


class AdminWindow(object):
    def setupUi(self, AdminWindow):
        AdminWindow.setObjectName("AdminWindow")
        AdminWindow.resize(743, 644)
        self.centralwidget = QtWidgets.QWidget(AdminWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(280, 10, 211, 41))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(10, 30, 721, 571))
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.books_groupBox = QtWidgets.QGroupBox(self.tab)
        self.books_groupBox.setGeometry(QtCore.QRect(10, 20, 691, 141))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(10)
        self.books_groupBox.setFont(font)
        self.books_groupBox.setObjectName("books_groupBox")
        self.title_label = QtWidgets.QLabel(self.books_groupBox)
        self.title_label.setGeometry(QtCore.QRect(10, 40, 41, 16))
        self.title_label.setObjectName("title_label")
        self.author_label = QtWidgets.QLabel(self.books_groupBox)
        self.author_label.setGeometry(QtCore.QRect(190, 40, 51, 16))
        self.author_label.setObjectName("author_label")
        self.stock_label = QtWidgets.QLabel(self.books_groupBox)
        self.stock_label.setGeometry(QtCore.QRect(460, 40, 41, 16))
        self.stock_label.setObjectName("stock_label")
        self.rack_label = QtWidgets.QLabel(self.books_groupBox)
        self.rack_label.setGeometry(QtCore.QRect(560, 40, 61, 16))
        self.rack_label.setObjectName("rack_label")
        self.id_label = QtWidgets.QLabel(self.books_groupBox)
        self.id_label.setGeometry(QtCore.QRect(380, 40, 21, 16))
        self.id_label.setObjectName("id_label")
        self.title_input = QtWidgets.QLineEdit(self.books_groupBox)
        self.title_input.setGeometry(QtCore.QRect(50, 40, 131, 20))
        self.title_input.setObjectName("title_input")
        self.author_input = QtWidgets.QLineEdit(self.books_groupBox)
        self.author_input.setGeometry(QtCore.QRect(250, 40, 101, 21))
        self.author_input.setObjectName("author_input")
        self.id_input = QtWidgets.QLineEdit(self.books_groupBox)
        self.id_input.setGeometry(QtCore.QRect(400, 40, 41, 21))
        self.id_input.setObjectName("id_input")
        self.stock_input = QtWidgets.QLineEdit(self.books_groupBox)
        self.stock_input.setGeometry(QtCore.QRect(510, 40, 31, 21))
        self.stock_input.setObjectName("stock_input")
        self.rack_input = QtWidgets.QLineEdit(self.books_groupBox)
        self.rack_input.setGeometry(QtCore.QRect(630, 40, 41, 21))
        self.rack_input.setObjectName("rack_input")
        self.viewAll_b = QtWidgets.QPushButton(self.books_groupBox)
        self.viewAll_b.setGeometry(QtCore.QRect(570, 90, 75, 31))
        self.viewAll_b.setObjectName("viewAll_b")
        self.addBook_b = QtWidgets.QPushButton(self.books_groupBox)
        self.addBook_b.setGeometry(QtCore.QRect(200, 90, 95, 31))
        self.addBook_b.setObjectName("addBook_b")
        self.editBook_b = QtWidgets.QPushButton(self.books_groupBox)
        self.editBook_b.setGeometry(QtCore.QRect(300, 90, 95, 31))
        self.editBook_b.setObjectName("editBook_b")
        self.removeBook_b = QtWidgets.QPushButton(self.books_groupBox)
        self.removeBook_b.setGeometry(QtCore.QRect(405, 90, 120, 31))
        self.removeBook_b.setObjectName("removeBook_b")
        self.search_b = QtWidgets.QPushButton(self.books_groupBox)
        self.search_b.setGeometry(QtCore.QRect(50, 90, 75, 31))
        self.search_b.setObjectName("search_b")
        self.books_list = QtWidgets.QListWidget(self.tab)
        self.books_list.setGeometry(QtCore.QRect(70, 220, 256, 311))
        self.books_list.setObjectName("books_list")
        self.list_label = QtWidgets.QLabel(self.tab)
        self.list_label.setGeometry(QtCore.QRect(80, 190, 111, 31))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.list_label.setFont(font)
        self.list_label.setObjectName("list_label")
        self.bookInfo = QtWidgets.QTextBrowser(self.tab)
        self.bookInfo.setGeometry(QtCore.QRect(370, 280, 300, 192))
        self.bookInfo.setObjectName("bookInfo")
        self.bookInfo_label = QtWidgets.QLabel(self.tab)
        self.bookInfo_label.setGeometry(QtCore.QRect(380, 250, 111, 31))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.bookInfo_label.setFont(font)
        self.bookInfo_label.setObjectName("bookInfo_label")
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.books_groupBox_2 = QtWidgets.QGroupBox(self.tab_2)
        self.books_groupBox_2.setGeometry(QtCore.QRect(10, 20, 691, 141))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(10)
        self.books_groupBox_2.setFont(font)
        self.books_groupBox_2.setObjectName("books_groupBox_2")
        self.name_label = QtWidgets.QLabel(self.books_groupBox_2)
        self.name_label.setGeometry(QtCore.QRect(10, 40, 41, 16))
        self.name_label.setObjectName("name_label")
        self.id_label_2 = QtWidgets.QLabel(self.books_groupBox_2)
        self.id_label_2.setGeometry(QtCore.QRect(10, 70, 21, 16))
        self.id_label_2.setObjectName("id_label_2")
        self.student_name = QtWidgets.QLineEdit(self.books_groupBox_2)
        self.student_name.setGeometry(QtCore.QRect(60, 40, 151, 20))
        self.student_name.setObjectName("student_name")
        self.student_id = QtWidgets.QLineEdit(self.books_groupBox_2)
        self.student_id.setGeometry(QtCore.QRect(60, 70, 151, 21))
        self.student_id.setObjectName("student_id")
        self.viewAll_b_2 = QtWidgets.QPushButton(self.books_groupBox_2)
        self.viewAll_b_2.setGeometry(QtCore.QRect(390, 50, 121, 31))
        self.viewAll_b_2.setObjectName("viewAll_b_2")
        self.addStudent_b = QtWidgets.QPushButton(self.books_groupBox_2)
        self.addStudent_b.setGeometry(QtCore.QRect(260, 50, 121, 31))
        self.addStudent_b.setObjectName("addStudent_b")
        #self.editStudent_b = QtWidgets.QPushButton(self.books_groupBox_2)
        #self.editStudent_b.setGeometry(QtCore.QRect(390, 50, 121, 31))
        #self.editStudent_b.setObjectName("editStudent_b")
        self.removeStudent_b = QtWidgets.QPushButton(self.books_groupBox_2)
        self.removeStudent_b.setGeometry(QtCore.QRect(520, 50, 141, 31))
        self.removeStudent_b.setObjectName("removeStudent_b")
        self.list_label_2 = QtWidgets.QLabel(self.tab_2)
        self.list_label_2.setGeometry(QtCore.QRect(100, 190, 121, 31))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.list_label_2.setFont(font)
        self.list_label_2.setObjectName("list_label_2")
        self.studentInfo_label = QtWidgets.QLabel(self.tab_2)
        self.studentInfo_label.setGeometry(QtCore.QRect(400, 250, 131, 31))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.studentInfo_label.setFont(font)
        self.studentInfo_label.setObjectName("studentInfo_label")
        self.students_list = QtWidgets.QListWidget(self.tab_2)
        self.students_list.setGeometry(QtCore.QRect(90, 220, 256, 311))
        self.students_list.setObjectName("students_list")
        self.student_info = QtWidgets.QTextBrowser(self.tab_2)
        self.student_info.setGeometry(QtCore.QRect(390, 280, 280, 192))
        self.student_info.setObjectName("student_info")
        self.tabWidget.addTab(self.tab_2, "")
        AdminWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(AdminWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 743, 21))
        self.menubar.setObjectName("menubar")
        self.menuFile = QtWidgets.QMenu(self.menubar)
        self.menuFile.setObjectName("menuFile")
        self.menuThemes = QtWidgets.QMenu(self.menubar)
        self.menuThemes.setObjectName("menuThemes")
        self.menuHelp = QtWidgets.QMenu(self.menubar)
        self.menuHelp.setObjectName("menuHelp")
        AdminWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(AdminWindow)
        self.statusbar.setObjectName("statusbar")
        AdminWindow.setStatusBar(self.statusbar)
        self.actionLibrarian = QtWidgets.QAction(AdminWindow)
        self.actionLibrarian.setObjectName("actionLibrarian")
        self.actionChange_Admin_Credentials = QtWidgets.QAction(AdminWindow)
        self.actionChange_Admin_Credentials.setObjectName("actionChange_Admin_Credentials")
        self.actionLogout = QtWidgets.QAction(AdminWindow)
        self.actionLogout.setObjectName("actionLogout")
        self.actionExit = QtWidgets.QAction(AdminWindow)
        self.actionExit.setObjectName("actionExit")
        self.actionBasic = QtWidgets.QAction(AdminWindow)
        self.actionBasic.setObjectName("actionBasic")
        self.actionAqua = QtWidgets.QAction(AdminWindow)
        self.actionAqua.setObjectName("actionAqua")
        self.actionCoral = QtWidgets.QAction(AdminWindow)
        self.actionCoral.setObjectName("actionCoral")
        self.actionAbout = QtWidgets.QAction(AdminWindow)
        self.actionAbout.setObjectName("actionAbout")
        self.menuFile.addAction(self.actionLibrarian)
        self.menuFile.addSeparator()
        self.menuFile.addAction(self.actionChange_Admin_Credentials)
        self.menuFile.addAction(self.actionLogout)
        self.menuFile.addSeparator()
        self.menuFile.addAction(self.actionExit)
        self.menuThemes.addAction(self.actionBasic)
        self.menuThemes.addAction(self.actionAqua)
        self.menuThemes.addAction(self.actionCoral)
        self.menuHelp.addAction(self.actionAbout)
        self.menubar.addAction(self.menuFile.menuAction())
        self.menubar.addAction(self.menuThemes.menuAction())
        self.menubar.addAction(self.menuHelp.menuAction())

        self.actionAqua.triggered.connect(self.change_to_blue)
        self.actionBasic.triggered.connect(self.change_back_to_basic)
        self.actionCoral.triggered.connect(self.change_to_coral)
        self.viewAll_b.clicked.connect(self.view_all_books)
        self.viewAll_b_2.clicked.connect(self.view_all_students)

        self.books_list.clicked.connect(self.show_book_description)
        self.students_list.clicked.connect(self.show_student_info)

        self.addBook_b.clicked.connect(self.add_book)
        self.addStudent_b.clicked.connect(self.add_student)

        self.removeBook_b.clicked.connect(self.remove_book)
        self.removeStudent_b.clicked.connect(self.remove_student)

        self.editBook_b.clicked.connect(self.edit_book)
        #self.editStudent_b.clicked.connect(self.edit_student)

        self.search_b.clicked.connect(self.search)
        self.actionLibrarian.triggered.connect(self.lend_return)
        self.actionLogout.triggered.connect(self.logout)
        self.actionChange_Admin_Credentials.triggered.connect(self.get_credentials)
        self.actionExit.triggered.connect(self.exit)
        self.actionAbout.triggered.connect(self.about)

        self.retranslateUi(AdminWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(AdminWindow)

    def retranslateUi(self, AdminWindow):
        _translate = QtCore.QCoreApplication.translate
        AdminWindow.setWindowTitle(_translate("AdminWindow", "Library Management System"))
        self.label.setText(_translate("AdminWindow", "Administrator Menu"))
        self.books_groupBox.setTitle(_translate("AdminWindow", "Books Database"))
        self.title_label.setText(_translate("AdminWindow", "Title"))
        self.author_label.setText(_translate("AdminWindow", "Author"))
        self.stock_label.setText(_translate("AdminWindow", "Stock"))
        self.rack_label.setText(_translate("AdminWindow", "Rack No"))
        self.id_label.setText(_translate("AdminWindow", "ID"))
        self.viewAll_b.setStatusTip(_translate("AdminWindow", "Display All Books in the Library\'s Collection"))
        self.viewAll_b.setText(_translate("AdminWindow", "View All"))
        self.addBook_b.setStatusTip(_translate("AdminWindow", "Add a Book to the Library\'s Collection"))
        self.addBook_b.setText(_translate("AdminWindow", "Add Book"))
        self.editBook_b.setStatusTip(_translate("AdminWindow", "Edit a Book in the Library\'s Collection"))
        self.editBook_b.setText(_translate("AdminWindow", "Edit Book"))
        self.removeBook_b.setStatusTip(_translate("AdminWindow", "Remove a Book from the Library\'s Collection"))
        self.removeBook_b.setText(_translate("AdminWindow", "Remove Book"))
        self.search_b.setStatusTip(_translate("AdminWindow", "Search for a Book by title in the Library\'s Collection"))
        self.search_b.setText(_translate("AdminWindow", "Search"))
        self.list_label.setText(_translate("AdminWindow", "Books List"))
        self.bookInfo_label.setText(_translate("AdminWindow", "Book\'s Info"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("AdminWindow", "Books Database"))
        self.books_groupBox_2.setTitle(_translate("AdminWindow", "Students Database"))
        self.name_label.setText(_translate("AdminWindow", "Name"))
        self.id_label_2.setText(_translate("AdminWindow", "ID"))
        self.viewAll_b_2.setStatusTip(_translate("AdminWindow", "Display All Students in the Library\'s Database"))
        self.viewAll_b_2.setText(_translate("AdminWindow", "View All"))
        self.addStudent_b.setStatusTip(_translate("AdminWindow", "Add a Student to the Database"))
        self.addStudent_b.setText(_translate("AdminWindow", "Add Student"))
        #self.editStudent_b.setStatusTip(_translate("AdminWindow", "Edit a Student in the Database"))
        #self.editStudent_b.setText(_translate("AdminWindow", "Edit Student"))
        self.removeStudent_b.setStatusTip(_translate("AdminWindow", "Remove a Student from the Database"))
        self.removeStudent_b.setText(_translate("AdminWindow", "Remove Student"))
        self.list_label_2.setText(_translate("AdminWindow", "Students List"))
        self.studentInfo_label.setText(_translate("AdminWindow", "Student\'s Info"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("AdminWindow", "Students Database"))
        self.menuFile.setTitle(_translate("AdminWindow", "File"))
        self.menuThemes.setTitle(_translate("AdminWindow", "Themes"))
        self.menuHelp.setTitle(_translate("AdminWindow", "Help"))
        self.actionLibrarian.setText(_translate("AdminWindow", "Lend/Return Books"))
        self.actionLibrarian.setStatusTip(_translate("AdminWindow", "Go to Lend/Borrow Books Menu"))
        self.actionLogout.setText(_translate("AdminWindow", "Logout"))
        self.actionLogout.setStatusTip(_translate("AdminWindow", "Go Back to the Login Menu"))
        self.actionLogout.setShortcut(_translate("AdminWindow", "Ctrl+Q"))
        self.actionChange_Admin_Credentials.setText(_translate("AdminWindow", "Change Admin Credentials"))
        self.actionChange_Admin_Credentials.setStatusTip(_translate("AdminWindow", "Change Administrator Credentials"))
        self.actionExit.setText(_translate("AdminWindow", "Exit"))
        self.actionExit.setStatusTip(_translate("AdminWindow", "Exit System"))
        self.actionExit.setShortcut(_translate("AdminWindow", "Esc"))
        self.actionBasic.setText(_translate("AdminWindow", "Basic"))
        self.actionBasic.setStatusTip(_translate("AdminWindow", "Basic Theme"))
        self.actionAqua.setText(_translate("AdminWindow", "Aqua"))
        self.actionAqua.setStatusTip(_translate("AdminWindow", "Blue-ish Theme"))
        self.actionCoral.setText(_translate("AdminWindow", "Coral"))
        self.actionCoral.setStatusTip(_translate("AdminWindow", "Pink-ish Theme"))
        self.actionAbout.setText(_translate("AdminWindow", "About"))
        self.actionAbout.setStatusTip(_translate("AdminWindow", "About the System"))
        self.actionAbout.setShortcut(_translate("AdminWindow", "Ctrl+H"))
        self.label.adjustSize()
        self.bookInfo_label.adjustSize()
        self.list_label.adjustSize()
        self.studentInfo_label.adjustSize()
        self.list_label_2.adjustSize()
        self.author_label.adjustSize()
        self.rack_label.adjustSize()
        self.stock_label.adjustSize()
        self.name_label.adjustSize()

    def about(self):
        msg = QMessageBox()
        msg.setWindowTitle("About")
        msg.setText("""The Library Management System LMS allows Admins to:
1. View the Library's entire Book Collection.
2. Search for any Book by Title.
3. Add to/Remove from/Edit the Library's Books Database.
4. Add to/Remove from/Edit the Library's Students Database.
5. Lend and Return Books.
6. Check for Lend Limits and Returns Late Fees.

Books can be borrowed for a period of one week, after which, the student must pay a late fee.""")

        x = msg.exec_()

    def logout(self):
        admin_menu.close()
        change_creds.close()
        lend_return_menu.close()
        login_screen.show()

    def change_back_to_basic(self):
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0,0,0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0,0,0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        admin_menu.setPalette(palette)
        change_creds.setPalette(palette)
        lend_return_menu.setPalette(palette)

    def change_to_coral(self):
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 85, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(183, 255, 251))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 170, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Highlight, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.HighlightedText, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.AlternateBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 85, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(183, 255, 251))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 170, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Highlight, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.HighlightedText, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.AlternateBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 120, 215))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Highlight, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.HighlightedText, brush)
        brush = QtGui.QBrush(QtGui.QColor(170, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.AlternateBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.PlaceholderText, brush)
        admin_menu.setPalette(palette)
        change_creds.setPalette(palette)
        lend_return_menu.setPalette(palette)

    def change_to_blue(self):
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 85, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 85, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 255, 253))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(127, 49, 191, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 85, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 85, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 255, 253))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(127, 49, 191, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 0, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 169, 200))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 0, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 0, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 255, 253))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 255, 253))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.PlaceholderText, brush)
        admin_menu.setPalette(palette)
        change_creds.setPalette(palette)
        lend_return_menu.setPalette(palette)

    def exit(self):
        admin_menu.close()
        change_creds.close()
        lend_return_menu.close()

    def view_all_books(self):
        self.id_input.clear()
        self.title_input.clear()
        self.author_input.clear()
        self.stock_input.clear()
        self.rack_input.clear()

        self.books_list.clear()
        self.bookInfo.clear()
        for book in LMS.get_available_books():
            self.books_list.addItem(book.get_title())

    def view_all_students(self):
        self.student_name.clear()
        self.student_id.clear()

        self.students_list.clear()
        self.student_info.clear()
        for student in LMS.get_students_list():
            self.students_list.addItem(student.get_name())

    def show_book_description(self):
        title = self.books_list.currentItem().text()

        for book in LMS.get_available_books():
            if title == book.get_title():
                text = book.get_info()

        self.bookInfo.setText(text)

    def show_student_info(self):
        name = self.students_list.currentItem().text()

        for student in LMS.get_students_list():
            if name == student.get_name():
                text = student.get_info()
                break

        self.student_info.setText(text)

    def add_book(self):
        title = self.title_input.text()
        author = self.author_input.text()
        id = self.id_input.text()
        stock = self.stock_input.text()
        rack = self.rack_input.text()

        if title == "" or author == "" or id == "":
            self.popup("Add Book", "Invalid Input.\nTo Add a Book, Add (at least) the following Info:\n1. ID\n2. Title\n3. Author", critical=True)
        else:
            LMS.add_book(id, title, author, stock, rack)
            self.popup("Add Book", "Book Successfully Added.")

            for book in LMS.get_available_books():
                if id == book.get_id():
                    max_row_add_book(book)
                    break

            self.view_all_books()

    def add_student(self):
        name = self.student_name.text()
        id = self.student_id.text()

        if name == "" or id == "":
            self.popup("Add Student", "Invalid Input.\nPlease Enter the Student's Name and ID.", critical=True)
        else:
            LMS.add_student(name, id)
            self.popup("Add Student", "Student Registered Successfully.")

            for student in LMS.get_students_list():
                if student.get_id() == id:
                    max_row_add_student(student)
                    break

            self.view_all_students()

    def remove_book(self):
        id = self.id_input.text()

        if id == "":
            self.popup("Remove Book", "Invalid Input.\nTo Remove a Book, You Must Specify its ID.", critical=True)
        else:
            book_found = False
            for book in LMS.get_available_books():
                if book.get_id() == id:
                    book_found = True
                    LMS.remove_book(id)
                    self.popup("Remove Book", "Book Successfully Removed.")
                    remove_book_from_db(id)
                    self.view_all_books()
                    break
            if not book_found:
                self.popup("Remove Book", "Book isn't in Library's Collection.\nYou Must Specify the Book's ID.", critical=True)

    def remove_student(self):
        id = self.student_id.text()

        if id == "":
            self.popup("Remove Student", "Invalid Input.\nTo Remove a Student, You Must Specify their ID.", critical=True)
        else:
            student_found = False
            for student in LMS.get_students_list():
                if student.get_id() == id:
                    student_found = True
                    LMS.remove_student(id)
                    self.popup("Remove Student", "Student Successfully Removed.")
                    remove_student_from_db(id)
                    self.view_all_students()
                    break
            if not student_found:
                self.popup("Remove Student", "Student isn't in Library's Database.\nYou Must Specify their ID.",
                           critical=True)

    def edit_book(self):
        title = self.title_input.text()
        author = self.author_input.text()
        id = self.id_input.text()
        stock = self.stock_input.text()
        rack = self.rack_input.text()

        if title == "" or author == "" or id == "":
            self.popup("Edit Book",
                       "Invalid Input.\nTo Edit a Book, Select it, and Add (at least) the following Info:\n1. ID\n2. Title\n3. Author",
                       critical=True)
        else:
            title_old = self.books_list.currentItem().text()

            for book in LMS.get_available_books():
                if title_old == book.get_title():
                    LMS.update_book(title_old, id, title, author, stock, rack)
                    self.popup("Edit Book", "Book Successfully Updated.")
                    edit_book_in_db(book, title_old)
                    self.view_all_books()
                    break

    def popup(self, title, message, critical=False):
        msg = QMessageBox()
        msg.setWindowTitle(title)
        msg.setText(message)
        if critical:
            msg.setIcon(QMessageBox.Critical)

        x = msg.exec_()

    def search(self):
        search_title = self.title_input.text()

        if search_title == "":
            self.popup("Book Search", "Please Enter a Title.", critical=True)
        else:
            books_found = LMS.search(search_title)

            if len(books_found) == 0:
                self.popup("Book Search", "Book isn't Available.", critical=True)

            else:
                self.books_list.clear()
                for book in books_found:
                    self.books_list.addItem(book.get_title())

    def get_credentials(self):
        change_creds.show()

    def lend_return(self):
        lend_return_menu.show()


class LendReturnWindow(object):
    def setupUi(self, LendReturnWindow):
        LendReturnWindow.setObjectName("LendReturnWindow")
        LendReturnWindow.resize(726, 553)
        self.centralwidget = QtWidgets.QWidget(LendReturnWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(270, 20, 201, 41))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.lend_groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.lend_groupBox.setGeometry(QtCore.QRect(20, 70, 681, 211))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lend_groupBox.setFont(font)
        self.lend_groupBox.setObjectName("lend_groupBox")
        self.book_id_label = QtWidgets.QLabel(self.lend_groupBox)
        self.book_id_label.setGeometry(QtCore.QRect(70, 60, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.book_id_label.setFont(font)
        self.book_id_label.setObjectName("book_id_label")
        self.student_id_label = QtWidgets.QLabel(self.lend_groupBox)
        self.student_id_label.setGeometry(QtCore.QRect(70, 90, 111, 16))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.student_id_label.setFont(font)
        self.student_id_label.setObjectName("student_id_label")
        self.book_id_input = QtWidgets.QLineEdit(self.lend_groupBox)
        self.book_id_input.setGeometry(QtCore.QRect(170, 60, 113, 20))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(10)
        font.setBold(False)
        font.setWeight(50)
        self.book_id_input.setFont(font)
        self.book_id_input.setObjectName("book_id_input")
        self.student_id_input = QtWidgets.QLineEdit(self.lend_groupBox)
        self.student_id_input.setGeometry(QtCore.QRect(170, 90, 113, 20))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(10)
        font.setBold(False)
        font.setWeight(50)
        self.student_id_input.setFont(font)
        self.student_id_input.setObjectName("student_id_input")
        self.lend_date_input = QtWidgets.QLineEdit(self.lend_groupBox)
        self.lend_date_input.setGeometry(QtCore.QRect(460, 60, 113, 20))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(10)
        font.setBold(False)
        font.setWeight(50)
        self.lend_date_input.setFont(font)
        self.lend_date_input.setObjectName("lend_date_input")
        self.lend_date_label = QtWidgets.QLabel(self.lend_groupBox)
        self.lend_date_label.setGeometry(QtCore.QRect(370, 60, 81, 16))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.lend_date_label.setFont(font)
        self.lend_date_label.setObjectName("lend_date_label")
        self.lend_book_b = QtWidgets.QPushButton(self.lend_groupBox)
        self.lend_book_b.setGeometry(QtCore.QRect(260, 150, 151, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.lend_book_b.setFont(font)
        self.lend_book_b.setObjectName("lend_book_b")
        self.return_groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.return_groupBox.setGeometry(QtCore.QRect(20, 300, 681, 211))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.return_groupBox.setFont(font)
        self.return_groupBox.setObjectName("return_groupBox")
        self.student_id_label_2 = QtWidgets.QLabel(self.return_groupBox)
        self.student_id_label_2.setGeometry(QtCore.QRect(70, 90, 111, 16))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.student_id_label_2.setFont(font)
        self.student_id_label_2.setObjectName("student_id_label_2")
        self.return_date_label = QtWidgets.QLabel(self.return_groupBox)
        self.return_date_label.setGeometry(QtCore.QRect(370, 60, 81, 16))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.return_date_label.setFont(font)
        self.return_date_label.setObjectName("return_date_label")
        self.return_date_input = QtWidgets.QLineEdit(self.return_groupBox)
        self.return_date_input.setGeometry(QtCore.QRect(470, 60, 113, 20))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(10)
        font.setBold(False)
        font.setWeight(50)
        self.return_date_input.setFont(font)
        self.return_date_input.setObjectName("return_date_input")
        self.book_id_input_2 = QtWidgets.QLineEdit(self.return_groupBox)
        self.book_id_input_2.setGeometry(QtCore.QRect(170, 60, 113, 20))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(10)
        font.setBold(False)
        font.setWeight(50)
        self.book_id_input_2.setFont(font)
        self.book_id_input_2.setObjectName("book_id_input_2")
        self.return_book_b = QtWidgets.QPushButton(self.return_groupBox)
        self.return_book_b.setGeometry(QtCore.QRect(260, 150, 151, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.return_book_b.setFont(font)
        self.return_book_b.setObjectName("return_book_b")
        self.student_id_input_2 = QtWidgets.QLineEdit(self.return_groupBox)
        self.student_id_input_2.setGeometry(QtCore.QRect(170, 90, 113, 20))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(10)
        font.setBold(False)
        font.setWeight(50)
        self.student_id_input_2.setFont(font)
        self.student_id_input_2.setObjectName("student_id_input_2")
        self.book_id_label_2 = QtWidgets.QLabel(self.return_groupBox)
        self.book_id_label_2.setGeometry(QtCore.QRect(70, 60, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.book_id_label_2.setFont(font)
        self.book_id_label_2.setObjectName("book_id_label_2")
        LendReturnWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(LendReturnWindow)
        self.statusbar.setObjectName("statusbar")
        LendReturnWindow.setStatusBar(self.statusbar)

        self.lend_book_b.clicked.connect(self.lend_book)
        self.return_book_b.clicked.connect(self.return_book)

        self.retranslateUi(LendReturnWindow)
        QtCore.QMetaObject.connectSlotsByName(LendReturnWindow)

    def retranslateUi(self, LendReturnWindow):
        _translate = QtCore.QCoreApplication.translate
        LendReturnWindow.setWindowTitle(_translate("LendReturnWindow", "Library Management System"))
        self.label.setText(_translate("LendReturnWindow", "Lend/Return Menu"))
        self.lend_groupBox.setTitle(_translate("LendReturnWindow", "Lend"))
        self.book_id_label.setText(_translate("LendReturnWindow", "Book\'s ID"))
        self.student_id_label.setText(_translate("LendReturnWindow", "Student\'s ID"))
        self.lend_date_label.setText(_translate("LendReturnWindow", "Lend Date"))
        self.lend_book_b.setStatusTip(_translate("LendReturnWindow", "Lend Selected Book After Checking Student\'s Book Limit"))
        self.lend_book_b.setText(_translate("LendReturnWindow", "Lend Book"))
        self.return_groupBox.setTitle(_translate("LendReturnWindow", "Return"))
        self.student_id_label_2.setText(_translate("LendReturnWindow", "Student\'s ID"))
        self.return_date_label.setText(_translate("LendReturnWindow", "Return Date"))
        self.return_book_b.setStatusTip(_translate("LendReturnWindow", "Return Selected Book After Checking Late Fees"))
        self.return_book_b.setText(_translate("LendReturnWindow", "Return Book"))
        self.book_id_label_2.setText(_translate("LendReturnWindow", "Book\'s Title"))
        self.label.adjustSize()
        self.return_date_label.adjustSize()
        self.return_date_input.setText("dd/mm/yyyy")
        self.lend_date_input.setText("dd/mm/yyyy")

    def popup(self, title, message, critical=False):
        msg = QMessageBox()
        msg.setWindowTitle(title)
        msg.setText(message)
        if critical:
            msg.setIcon(QMessageBox.Critical)

        x = msg.exec_()

    def lend_book(self):
        book_id = self.book_id_input.text()
        student_id = self.student_id_input.text()
        date = self.lend_date_input.text()

        if book_id == "" or student_id == "" or date == "":
            self.popup("Lend Book", "Invalid Input. All Fields Must be Filled", critical=True)

        else:

            # Check if book id is in database
            book_found = False
            for book in LMS.get_available_books():
                if book.get_id() == book_id:

                    # Book Found
                    book_found = True

                    # Not in Stock
                    if book.get_stock() == '0':
                        self.popup("Lend Book", "Sorry the Requested Book is Currently Out of Stock", critical=True)
                        break

                    # In Stock
                    else:

                        # Check if Student is in database
                        student_found = False
                        for student in LMS.get_students_list():
                            if student.get_id().lower() == student_id.lower():

                                # Student Found
                                student_found = True

                                # Check Limit
                                if student.has_passed_book_limit() is False:

                                    # Check Limit Passed
                                    student.borrow_book(book.get_title(), date)
                                    remove_student_from_db(student_id)
                                    max_row_add_student(student)
                                    book.add_stock(-1)
                                    edit_book_in_db(book, book.get_title())
                                    self.popup("Lend Book", "Book is Successfully Lent")
                                    self.book_id_input.clear()
                                    self.student_id_input.clear()
                                    self.lend_date_input.clear()
                                    AdminWindow.view_all_books(admin_menu_ui_instance)
                                    AdminWindow.view_all_students(admin_menu_ui_instance)
                                    break

                                # Check Limit Failed
                                else:
                                    self.popup("Lend Book", "Book Limit (3) Exceeded.", critical=True)
                                    break

                        # Student not found
                        if not student_found:
                            self.popup("Lend Book", "Student's ID Not Found in Database", critical=True)
                            break

            # Book not found
            if not book_found:
                self.popup("Lend Book", "Book's ID Not Found in Database", critical=True)

    def return_book(self):
        book_title = self.book_id_input_2.text()
        student_id = self.student_id_input_2.text()
        return_date = self.return_date_input.text()

        if book_title == "" or student_id == "" or return_date == "":
            self.popup("Return Book", "Invalid Input. All Fields Must be Filled", critical=True)

        else:

            bb = ""
            b = Book
            s = Student

            # Check if book is in database
            book_found = False
            for book in LMS.get_available_books():
                if book.get_title().lower() == book_title.lower():
                    book_found = True
                    b = book
                    break

            # Check if student is in database
            student_found = False
            for student in LMS.get_students_list():
                if student.get_id().lower() == student_id.lower():
                    student_found = True
                    s = student
                    break

            # Check if student borrowed book
            sborrowedb = False
            if student_found and book_found:
                for borrowed_book in s.get_borrowed_books_list():
                    if borrowed_book.lower() == book_title.lower():
                        sborrowedb = True
                        bb = borrowed_book
                        break

            if student_found and book_found and sborrowedb:
                borrow_date = s.get_borrow_date(book_title)

                if not check_return_date_validity(borrow_date, return_date):
                    self.popup("Return Book", "Invalid Return Date!", critical=True)
                else:

                    b.add_stock(1)
                    s.return_book(bb)
                    remove_student_from_db(student_id)
                    max_row_add_student(s)
                    self.popup("Return Book", "Book is Successfully Returned")
                    self.book_id_input_2.clear()
                    self.student_id_input_2.clear()
                    self.return_date_input.clear()
                    AdminWindow.view_all_books(admin_menu_ui_instance)
                    AdminWindow.view_all_students(admin_menu_ui_instance)

                    fee = LMS.get_late_fee(borrow_date, return_date)
                    if fee > 1:
                        self.popup("Return Book", f"Late Book Return!\n"
                                                  f"Borrow Date: {borrow_date}\n"
                                                  f"Return Date: {return_date}\n"
                                                  f"Late Fee: {fee}LE", critical=True)

            if not book_found:
                self.popup("Return Book", "Book Not Found in Database.\nPlease Enter a Valid Book ID", critical=True)

            if not student_found:
                self.popup("Return Book", "Student ID Entered Not Registered in Database", critical=True)

            if student_found and book_found and not sborrowedb:
                self.popup("Return Book", "Student Entered did not borrow the Book Entered", critical=True)


wb = openpyxl.load_workbook("Database.xlsx")
books_db = wb["Books Database"]
students_db = wb["Students Database"]
admin_db = wb["Admin Database"]

students = get_students_from_db()
books = get_books_from_db()
admin_info = get_admin_info_from_db()

LMS = Library(books, students, admin_info)


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)

    login_screen = QtWidgets.QMainWindow()
    students_menu = QtWidgets.QMainWindow()
    admin_menu = QtWidgets.QMainWindow()
    change_creds = QtWidgets.QMainWindow()
    lend_return_menu = QtWidgets.QMainWindow()

    login_screen_ui_instance = LoginScreen()
    login_screen_ui_instance.setupUi(login_screen)

    students_menu_ui_instance = StudentsMenu()
    students_menu_ui_instance.setupUi(students_menu)

    admin_menu_ui_instance = AdminWindow()
    admin_menu_ui_instance.setupUi(admin_menu)

    change_creds_ui_instance = ChangeCredsWindow()
    change_creds_ui_instance.setupUi(change_creds)

    lend_return_menu_ui_instance = LendReturnWindow()
    lend_return_menu_ui_instance.setupUi(lend_return_menu)

    login_screen.show()
    sys.exit(app.exec_())
